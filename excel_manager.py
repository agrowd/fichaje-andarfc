# -*- coding: utf-8 -*-
"""
excel_manager.py — Excel file scanner, photo extractor, and exporter.
Handles reading player data and embedded images from Excel files,
and writing photos back for export.
"""

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
import zipfile
import os
import io
import re
from PIL import Image

# Directory containing the Excel files
EXCEL_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "LISTAS DE BUENA FE AFA SOMOS TODOS 2026"
)

EXPORT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "exportados"
)

# Minimum acceptable photo dimensions
MIN_PHOTO_WIDTH = 150
MIN_PHOTO_HEIGHT = 150


def get_excel_files():
    """List all Excel files in the data directory."""
    if not os.path.isdir(EXCEL_DIR):
        return []
    return [
        f for f in sorted(os.listdir(EXCEL_DIR))
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')
    ]


def extract_team_name(filename):
    """Extract team name from filename (remove .xlsx and trailing numbers)."""
    name = os.path.splitext(filename)[0]
    # Clean up: "BOCA 1" -> "BOCA 1", "ALTE. BROWN" -> "ALTE. BROWN"
    return name.strip()


def scan_excel(filepath):
    """
    Scan an Excel file and extract all player data and photos.
    
    Returns:
        dict with keys:
            - filename: str
            - team_name: str
            - sheet_name: str
            - players: list of dicts with player info
            - images_by_row: dict mapping row_index -> image_bytes
            - total_images: int
            - has_unlinked_images: bool (images in zip but not in drawing)
    """
    filename = os.path.basename(filepath)
    result = {
        "filename": filename,
        "team_name": extract_team_name(filename),
        "sheet_name": "Data",
        "players": [],
        "images_by_row": {},
        "total_images": 0,
        "has_unlinked_images": False,
    }

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        print(f"[EXCEL] Error loading {filename}: {e}")
        return result

    # Find the data sheet
    if 'Data' in wb.sheetnames:
        ws = wb['Data']
        result["sheet_name"] = "Data"
    else:
        ws = wb[wb.sheetnames[0]]
        result["sheet_name"] = wb.sheetnames[0]

    # Extract players (rows 4+)
    for row_idx in range(4, ws.max_row + 1):
        name_val = ws.cell(row=row_idx, column=3).value  # Column C
        surname_val = ws.cell(row=row_idx, column=4).value  # Column D

        if not name_val and not surname_val:
            continue

        player = {
            "row": row_idx,
            "competition": _clean_value(ws.cell(row=row_idx, column=1).value),
            "team": _clean_value(ws.cell(row=row_idx, column=2).value),
            "name": _clean_value(name_val),
            "surname": _clean_value(surname_val),
            "email": _clean_value(ws.cell(row=row_idx, column=5).value),
            "dni": _clean_dni(ws.cell(row=row_idx, column=6).value),
        }
        result["players"].append(player)

    # Extract images anchored via drawing.xml (openpyxl detects these)
    for img in ws._images:
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from'):
                row_num = anchor._from.row + 1  # Convert 0-indexed to 1-indexed
                col_num = anchor._from.col
            else:
                continue

            # Read image data
            img_data = _read_image_data(img)
            if img_data:
                result["images_by_row"][row_num] = img_data
                result["total_images"] += 1
        except Exception as e:
            print(f"[EXCEL] Error reading image in {filename}: {e}")

    # Check for unlinked images (in xl/media but no drawing.xml)
    if result["total_images"] == 0:
        try:
            unlinked = _extract_unlinked_images(filepath)
            if unlinked:
                result["has_unlinked_images"] = True
                result["total_images"] = len(unlinked)
                # Try to associate by order with player rows
                for i, (img_name, img_data) in enumerate(unlinked):
                    if i < len(result["players"]):
                        player_row = result["players"][i]["row"]
                        result["images_by_row"][player_row] = img_data
        except Exception as e:
            print(f"[EXCEL] Error extracting unlinked images from {filename}: {e}")

    wb.close()
    return result


def scan_all_excels():
    """Scan all Excel files and return a list of results."""
    files = get_excel_files()
    results = []

    for fname in files:
        fpath = os.path.join(EXCEL_DIR, fname)
        print(f"[EXCEL] Scanning: {fname}")
        result = scan_excel(fpath)
        results.append(result)
        print(f"  Players: {len(result['players'])}, "
              f"Images: {result['total_images']}, "
              f"Unlinked: {result['has_unlinked_images']}")

    return results


def assess_photo_quality(photo_data):
    """
    Assess the quality of a photo. Returns a status string:
    - 'ok': Photo meets quality standards
    - 'bad': Photo is too small, corrupted, or poor quality
    - 'pending_review': Photo exists but needs manual review
    """
    if not photo_data:
        return "missing"

    try:
        img = Image.open(io.BytesIO(photo_data))
        width, height = img.size

        # Check minimum dimensions
        if width < MIN_PHOTO_WIDTH or height < MIN_PHOTO_HEIGHT:
            return "bad"

        # Check if image is mostly one color (potentially a placeholder)
        if _is_blank_image(img):
            return "bad"

        # Check aspect ratio (should be roughly portrait for ID photos)
        # But be lenient — some might be landscape crops
        return "ok"

    except Exception:
        return "bad"


def export_excel_with_photos(team_filename, players_with_photos):
    """
    Create a new Excel file with photos embedded in column G.
    Replicates the exact format of reference files (BOCA 1, LANUS 1, SAN LORENZO 1, etc.):
    - oneCellAnchor in column G (index 6), row offset 0, col offset 0
    - Column G width: ~25 units
    - Row heights: ~95pt per player row with photo
    - Photos sized to fit: ~130x170 pixels (portrait, carnet-style)
    
    Args:
        team_filename: Original filename (e.g., "BOCA 1.xlsx")
        players_with_photos: list of dicts with keys:
            - excel_row: int
            - photo_data: bytes (JPEG/PNG)
    
    Returns:
        Path to the exported file.
    """
    os.makedirs(EXPORT_DIR, exist_ok=True)

    source_path = os.path.join(EXCEL_DIR, team_filename)
    export_path = os.path.join(EXPORT_DIR, team_filename)

    # Load the original workbook
    wb = openpyxl.load_workbook(source_path)
    ws = wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[0]]

    # Remove existing images (we'll re-add all of them)
    ws._images = []

    # Match reference format: Column G width ~25 units (like LANUS 1 / SAN LORENZO 1)
    ws.column_dimensions['G'].width = 25

    for player in players_with_photos:
        if not player.get("photo_data"):
            continue

        row = player["excel_row"]

        try:
            # Match reference format: Row height ~95pt (refs range from 70-130pt)
            ws.row_dimensions[row].height = 95

            # Prepare the photo: resize to carnet/ficha proportions
            img_stream = io.BytesIO(player["photo_data"])
            pil_img = Image.open(img_stream)

            # Target size: portrait ratio ~3:4, fitting within the cell
            # Reference sizes in EMUs: ~500000-1000000 cx, ~700000-1100000 cy
            # That translates roughly to 130x170 pixels
            target_w, target_h = 130, 170

            # Maintain aspect ratio while fitting within target
            orig_w, orig_h = pil_img.size
            ratio = min(target_w / orig_w, target_h / orig_h)
            new_w = int(orig_w * ratio)
            new_h = int(orig_h * ratio)

            pil_img = pil_img.resize((new_w, new_h), Image.Resampling.LANCZOS)

            # Convert to RGB JPEG (like reference files)
            if pil_img.mode in ("RGBA", "P"):
                pil_img = pil_img.convert("RGB")

            resized_stream = io.BytesIO()
            pil_img.save(resized_stream, format="JPEG", quality=90)
            resized_stream.seek(0)

            # Create Excel image with proper dimensions
            xl_img = XlImage(resized_stream)
            xl_img.width = new_w
            xl_img.height = new_h

            # Anchor to column G at the player's row (like reference files)
            # Reference: col=6, colOff=0, row=N, rowOff=0 (SAN LORENZO format)
            cell_ref = f"G{row}"
            ws.add_image(xl_img, cell_ref)

        except Exception as e:
            print(f"[EXPORT] Error embedding photo at row {row}: {e}")

    wb.save(export_path)
    wb.close()

    print(f"[EXPORT] Saved: {export_path}")
    return export_path


def _read_image_data(xl_image):
    """Extract raw bytes from an openpyxl Image object."""
    try:
        if hasattr(xl_image, '_data'):
            return xl_image._data()
        elif hasattr(xl_image, 'ref'):
            ref = xl_image.ref
            if hasattr(ref, 'read'):
                ref.seek(0)
                return ref.read()
            elif isinstance(ref, (bytes, bytearray)):
                return bytes(ref)
        # Try to get from the internal blob
        if hasattr(xl_image, '_data'):
            return xl_image._data
        return None
    except Exception:
        return None


def _extract_unlinked_images(filepath):
    """
    Extract images from xl/media/ that are not linked via drawing.xml.
    Returns list of (filename, image_bytes) sorted by filename.
    """
    images = []
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            # Check if there's a drawing file — if yes, images ARE linked
            drawings = [f for f in z.namelist() if 'drawing' in f.lower() and f.endswith('.xml')]
            if drawings:
                return []  # Images are linked, not "unlinked"

            media_files = sorted([
                f for f in z.namelist()
                if f.startswith('xl/media/') and any(
                    f.lower().endswith(ext) for ext in ('.jpg', '.jpeg', '.png', '.gif', '.bmp')
                )
            ])

            for mf in media_files:
                data = z.read(mf)
                images.append((os.path.basename(mf), data))
    except Exception:
        pass

    return images


def _clean_value(val):
    """Clean a cell value to a string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_dni(val):
    """Clean a DNI value — remove dots, prefixes, normalize."""
    if val is None:
        return None
    s = str(val).strip()
    # Remove common prefixes
    s = re.sub(r'^dni:', '', s, flags=re.IGNORECASE)
    # Remove dots and spaces
    s = s.replace('.', '').replace(' ', '')
    # Remove .0 from float conversion
    if s.endswith('.0'):
        s = s[:-2]
    # Remove non-numeric chars except for special cases
    if s and not s.replace('-', '').isdigit():
        # Keep as-is if it's something like "#VALUE!"
        pass
    return s if s else None


def _is_blank_image(img, threshold=0.95):
    """Check if an image is mostly one solid color (blank/placeholder)."""
    try:
        # Sample pixels
        small = img.resize((20, 20), Image.Resampling.NEAREST)
        if small.mode != 'RGB':
            small = small.convert('RGB')
        pixels = list(small.getdata())

        if not pixels:
            return True

        # Check if most pixels are very similar
        first = pixels[0]
        similar_count = sum(
            1 for p in pixels
            if abs(p[0] - first[0]) < 20 and abs(p[1] - first[1]) < 20 and abs(p[2] - first[2]) < 20
        )

        return (similar_count / len(pixels)) > threshold
    except Exception:
        return False


if __name__ == "__main__":
    print(f"[EXCEL] Data dir: {EXCEL_DIR}")
    print(f"[EXCEL] Export dir: {EXPORT_DIR}")
    print(f"[EXCEL] Files found: {len(get_excel_files())}")
    print()
    results = scan_all_excels()

    print(f"\n{'='*50}")
    print(f"TOTAL FILES: {len(results)}")
    total_players = sum(len(r['players']) for r in results)
    total_images = sum(r['total_images'] for r in results)
    print(f"TOTAL PLAYERS: {total_players}")
    print(f"TOTAL IMAGES: {total_images}")
    print(f"MISSING: {total_players - total_images}")
