import openpyxl
import os

template_path = os.path.join("LISTAS DE BUENA FE AFA SOMOS TODOS 2026", "TEST.xlsx")
wb = openpyxl.load_workbook(template_path)
ws = wb.active

# Clear all data from row 4 down, except column headers
for row in range(4, 50):
    for col in range(1, 10):
        ws.cell(row=row, column=col).value = None

# Add our 3 test players exactly like the DB has them
test_players = [
    {"row": 4, "name": "Juan Carlos Pérez", "dni": "40123456", "email": "juan@test.com"},
    {"row": 5, "name": "María González", "dni": "41234567", "email": "maria@test.com"},
    {"row": 6, "name": "Pedro Rodríguez", "dni": "42345678", "email": "pedro@test.com"}
]

for p in test_players:
    r = p["row"]
    ws.cell(row=r, column=1).value = p["name"]
    ws.cell(row=r, column=2).value = "DNI"
    ws.cell(row=r, column=3).value = p["dni"]
    ws.cell(row=r, column=8).value = p["email"]

wb.save(template_path)
print("TEST.xlsx has been cleaned and populated with only the 3 test players.")
